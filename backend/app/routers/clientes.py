import io
import re
import unicodedata
import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user
from app.models.cliente import Cliente
from app.models.processo import Processo
from app.models.usuario import Usuario
from app.schemas.cliente import ClienteCreate, ClienteOut, ClienteUpdate, ClienteWithProcessos
from app.schemas.processo import ProcessoOut

UPLOADS_DIR = Path("/app/uploads/clientes")
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

router = APIRouter(prefix="/clientes", tags=["clientes"])


def _slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = re.sub(r"[^\w\s-]", "", text.lower())
    return re.sub(r"[-\s]+", "-", text).strip("-")


def _gerar_projeto(nome: str) -> tuple[str, str]:
    projeto_nome = f"{nome} — Contexto Jurídico"
    worktree_nome = f"cliente-{_slugify(nome)}"
    return projeto_nome, worktree_nome


def _pasta_cliente(cliente_id: uuid.UUID) -> Path:
    p = UPLOADS_DIR / str(cliente_id)
    p.mkdir(parents=True, exist_ok=True)
    return p


# ── CRUD ──────────────────────────────────────────────────────────────────────

@router.get("/", response_model=list[ClienteOut])
def listar_clientes(db: Session = Depends(get_db)):
    return db.query(Cliente).order_by(Cliente.nome).all()


@router.post("/", response_model=ClienteOut, status_code=status.HTTP_201_CREATED)
def criar_cliente(data: ClienteCreate, db: Session = Depends(get_db)):
    projeto_nome, worktree_nome = _gerar_projeto(data.nome)
    cliente = Cliente(**data.model_dump(), projeto_nome=projeto_nome, worktree_nome=worktree_nome)
    db.add(cliente)
    db.commit()
    db.refresh(cliente)
    try:
        from app.services.google_drive import ensure_cliente_folder
        ensure_cliente_folder(cliente.nome)
    except Exception:
        pass
    return cliente


# ── Pastas de cliente no Drive: órfãos + duplicatas (admin) ────────────────────
# Rotas fixas ANTES de /{cliente_id} — se ficassem depois, "admin" seria
# interpretado como cliente_id pelo FastAPI (rota dinâmica engole rota fixa).

class MesclarPastasBody(BaseModel):
    folder_ids: list[str]
    canonical_id: str | None = None


@router.get("/admin/pastas-drive/duplicatas")
def listar_duplicatas_drive(current: Usuario = Depends(get_current_user)):
    """Detecta (sem mesclar) pastas de cliente duplicadas na raiz do Drive."""
    if current.role != "super_admin":
        raise HTTPException(status_code=403, detail="Apenas super admin.")
    from app.services import drive_folder_heal as heal
    return {"duplicatas": heal.escanear_duplicatas_raiz()}


@router.get("/admin/pastas-drive/orfaos")
def listar_orfaos_drive(current: Usuario = Depends(get_current_user)):
    """Lista clientes cujo drive_folder_id aponta para pasta apagada/lixeira."""
    if current.role != "super_admin":
        raise HTTPException(status_code=403, detail="Apenas super admin.")
    from app.services import drive_folder_heal as heal
    return {"orfaos": heal.escanear_orfaos()}


@router.post("/admin/pastas-drive/mesclar")
def mesclar_pastas_drive(body: MesclarPastasBody, current: Usuario = Depends(get_current_user)):
    """Mescla um grupo de pastas duplicadas: move conteúdo das extras para a
    canônica e joga as extras vazias na lixeira. Nunca apaga pasta com
    conteúdo remanescente."""
    if current.role != "super_admin":
        raise HTTPException(status_code=403, detail="Apenas super admin.")
    from app.services import drive_folder_heal as heal
    return heal.mesclar_cluster(body.folder_ids, body.canonical_id)


@router.get("/{cliente_id}", response_model=ClienteWithProcessos)
def obter_cliente(cliente_id: uuid.UUID, db: Session = Depends(get_db)):
    from app.routers.processos import _processo_to_out
    from sqlalchemy import inspect as _sa_inspect

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    # Build response manually so ProcessoOut.clientes_litisconsorcio is correctly serialized
    cliente_cols = {c.key: getattr(cliente, c.key) for c in _sa_inspect(Cliente).column_attrs}
    processos_out = [_processo_to_out(p, db) for p in cliente.processos]
    return {**cliente_cols, "processos": processos_out}


@router.patch("/{cliente_id}", response_model=ClienteOut)
def atualizar_cliente(
    cliente_id: uuid.UUID, data: ClienteUpdate, db: Session = Depends(get_db)
):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    nome_antigo = cliente.nome
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(cliente, field, value)
    # Atualiza projeto se nome mudou
    if data.nome:
        cliente.projeto_nome, cliente.worktree_nome = _gerar_projeto(data.nome)
    db.commit()
    db.refresh(cliente)
    # Renomear a pasta-raiz no Drive quando o nome muda — mantém uma única pasta
    # por cliente (sem duplicar) e o Drive legível. O vínculo é pelo id gravado.
    if data.nome and (cliente.nome or "").strip() != (nome_antigo or "").strip():
        try:
            from app.services.google_drive import renomear_pasta_cliente
            fid = renomear_pasta_cliente(
                cliente.nome, folder_id=cliente.drive_folder_id, nome_antigo=nome_antigo
            )
            if fid and fid != cliente.drive_folder_id:
                cliente.drive_folder_id = fid
                db.commit()
                db.refresh(cliente)
        except Exception:
            pass
    return cliente


@router.delete("/{cliente_id}", status_code=status.HTTP_204_NO_CONTENT)
def deletar_cliente(cliente_id: uuid.UUID, db: Session = Depends(get_db)):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    db.delete(cliente)
    db.commit()


# ── Documentos PDF ────────────────────────────────────────────────────────────

@router.post("/{cliente_id}/upload-pdf")
async def upload_pdf(
    cliente_id: uuid.UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Apenas arquivos PDF são aceitos")

    pasta = _pasta_cliente(cliente_id)
    dest = pasta / file.filename
    content = await file.read()
    dest.write_bytes(content)
    drive_link = None
    try:
        from app.services.google_drive import upload_arquivo
        drive_link = upload_arquivo(
            content,
            file.filename,
            cliente.nome,
            "Uploads",
            file.content_type or "application/pdf",
        )
    except Exception:
        pass
    return {"filename": file.filename, "size": len(content), "drive_link": drive_link}


@router.get("/{cliente_id}/documentos")
def listar_documentos(cliente_id: uuid.UUID, db: Session = Depends(get_db)):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    # Arquivos locais do servidor (apoio imediato para IA) + Drive (fonte persistente)
    pasta = UPLOADS_DIR / str(cliente_id)
    server_docs = []
    if pasta.exists():
        server_docs = [
            {"filename": f.name, "size": f.stat().st_size, "subpasta": None, "source": "local"}
            for f in sorted(pasta.iterdir())
            if f.suffix.lower() in (".pdf", ".md", ".txt")
        ]

    try:
        from app.services.google_drive import listar_arquivos
        drive_docs = listar_arquivos(cliente.nome, "Uploads")
    except Exception:
        drive_docs = []

    # Merge e deduplica por nome de arquivo, preferindo o link persistente do Drive.
    nomes_server = {d["filename"] for d in server_docs}
    merged = list(drive_docs)
    nomes_drive = {d["filename"] for d in drive_docs}
    for doc in server_docs:
        if doc["filename"] not in nomes_drive:
            merged.append(doc)

    return merged


@router.delete("/{cliente_id}/documentos/{filename}")
def remover_documento(
    cliente_id: uuid.UUID, filename: str, db: Session = Depends(get_db)
):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    removed = False
    f = UPLOADS_DIR / str(cliente_id) / filename
    if f.exists():
        f.unlink()
        removed = True
    try:
        from app.services.google_drive import deletar_arquivo
        removed = deletar_arquivo(cliente.nome, "Uploads", filename) or removed
    except Exception:
        pass
    if not removed:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")
    return {"ok": True}


# ── Batch processos via XLS ──────────────────────────────────────────────────

COLUNAS_TEMPLATE = [
    "numero_cnj", "vara", "comarca", "estado", "tribunal",
    "materia", "fase", "status", "polo", "objeto",
]

ESTADOS_VALIDOS = {"ES", "SP", "AM", "RJ", "outro"}
FASES_VALIDAS = {"conhecimento", "recursal", "execucao", "cumprimento_sentenca", "outro", ""}
STATUS_VALIDOS = {"ativo", "suspenso", "arquivado", "encerrado", ""}
POLOS_VALIDOS = {
    "autor", "reu", "litisconsorte", "assistente", "opoente",
    "interveniente", "perito", "avaliador", "interessado", "outro", "",
}


@router.get("/{cliente_id}/processos/template-xls")
def download_template_processos(cliente_id: uuid.UUID, db: Session = Depends(get_db)):
    """Retorna um arquivo Excel com o cabeçalho padrão para importação em lote."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos"

    # Cabeçalho
    headers = COLUNAS_TEMPLATE
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1D1E20", end_color="1D1E20", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Linha de exemplo
    exemplo = [
        "0000000-00.0000.0.00.0000",  # numero_cnj
        "1ª Vara Cível",              # vara
        "Vitória",                    # comarca
        "ES",                         # estado (ES/SP/AM/RJ/outro)
        "TJES",                       # tribunal
        "Direito Civil",              # materia
        "conhecimento",               # fase
        "ativo",                      # status
        "autor",                      # polo
        "Ação de indenização",        # objeto
    ]
    for col, val in enumerate(exemplo, 1):
        ws.cell(row=2, column=col, value=val)

    # Ajusta largura das colunas
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=processos_{cliente_id}.xlsx"},
    )


@router.post("/{cliente_id}/processos/batch", response_model=list[ProcessoOut])
async def importar_processos_batch(
    cliente_id: uuid.UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Importa processos em lote a partir de um arquivo Excel."""
    import openpyxl

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Apenas arquivos Excel (.xlsx) são aceitos")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo Excel inválido")

    # Mapeia cabeçalhos
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "numero_cnj" not in headers:
        raise HTTPException(status_code=400, detail="Coluna 'numero_cnj' não encontrada — use o template")

    criados = []
    erros = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        values = {headers[i]: str(v).strip() if v is not None else "" for i, v in enumerate(row) if i < len(headers)}

        cnj = values.get("numero_cnj", "")
        if not cnj:
            erros.append(f"Linha {row_idx}: numero_cnj vazio")
            continue

        # Verifica duplicata
        if db.query(Processo).filter(Processo.numero_cnj == cnj).first():
            erros.append(f"Linha {row_idx}: {cnj} já existe")
            continue

        estado_raw = values.get("estado", "ES")
        if estado_raw not in ESTADOS_VALIDOS:
            estado_raw = "outro"

        fase_raw = values.get("fase", "") or None
        if fase_raw and fase_raw not in FASES_VALIDAS:
            fase_raw = "outro"

        status_raw = values.get("status", "ativo") or "ativo"
        if status_raw not in STATUS_VALIDOS:
            status_raw = "ativo"

        polo_raw = values.get("polo", "") or None
        if polo_raw and polo_raw not in POLOS_VALIDOS:
            polo_raw = "outro"

        proc = Processo(
            cliente_id=cliente_id,
            numero_cnj=cnj,
            vara=values.get("vara") or None,
            comarca=values.get("comarca") or None,
            estado=estado_raw,
            tribunal=values.get("tribunal") or None,
            materia=values.get("materia") or None,
            fase=fase_raw,
            status=status_raw,
            polo=polo_raw,
            objeto=values.get("objeto") or None,
        )
        db.add(proc)
        criados.append(proc)

    if erros and not criados:
        raise HTTPException(status_code=422, detail="; ".join(erros))

    db.commit()
    for p in criados:
        db.refresh(p)

    return criados


# ── Proposta PDF ─────────────────────────────────────────────────────────────

class PropostaRequest(BaseModel):
    projeto_tipo: str = "PPS"
    valor: str = ""
    condicao_pagamento: str = ""
    intro_texto: str = ""
    anamnese_resumo: str = ""
    blocos_extras: dict | None = None  # {bloco_nome: [item, ...]}


@router.post("/{cliente_id}/proposta")
def gerar_proposta(
    cliente_id: uuid.UUID,
    body: PropostaRequest,
    db: Session = Depends(get_db),
):
    from datetime import date
    from app.services.proposta_pdf import gerar_proposta as _gerar

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    pdf_bytes = _gerar(
        cliente_nome=cliente.nome,
        projeto_tipo=body.projeto_tipo,
        valor=body.valor,
        condicao_pagamento=body.condicao_pagamento,
        intro_texto=body.intro_texto,
        anamnese_resumo=body.anamnese_resumo,
        data_proposta=date.today(),
        blocos_extras=body.blocos_extras,
    )

    # Salva proposta em uploads/clientes/{id}/
    pasta = _pasta_cliente(cliente_id)
    from datetime import datetime
    nome_arquivo = f"Proposta_{body.projeto_tipo}_{datetime.today().strftime('%Y%m%d')}.pdf"
    (pasta / nome_arquivo).write_bytes(pdf_bytes)

    try:
        from app.services.google_drive import upload_arquivo
        upload_arquivo(pdf_bytes, nome_arquivo, cliente.nome, "Propostas")
    except Exception:
        pass

    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


# ── Chat IA ───────────────────────────────────────────────────────────────────

class ChatClienteRequest(BaseModel):
    pergunta: str
    historico: list[dict] = []
    modelo: str = "claude"


@router.post("/{cliente_id}/chat")
def chat_cliente(
    cliente_id: uuid.UUID,
    body: ChatClienteRequest,
    db: Session = Depends(get_db),
    current: Usuario = Depends(get_current_user),
):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    from app.services.contexto_service import montar_contexto_cliente
    contexto = montar_contexto_cliente(db, cliente, current)

    from app.services.ia_cliente import chat
    resposta = chat(
        modelo=body.modelo,
        pergunta=body.pergunta,
        historico=body.historico,
        cliente_id=str(cliente_id),
        contexto=contexto,
        nome_cliente=cliente.nome,
    )
    return {"resposta": resposta}


@router.get("/{cliente_id}/contexto")
def obter_contexto_cliente(
    cliente_id: uuid.UUID,
    db: Session = Depends(get_db),
    current: Usuario = Depends(get_current_user),
):
    """Retorna o texto exato que o chat IA recebe como contexto deste cliente."""
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    from app.services.contexto_service import montar_contexto_cliente
    return {"contexto": montar_contexto_cliente(db, cliente, current)}


@router.post("/{cliente_id}/emails/classificar-pendentes")
def classificar_emails_pendentes(
    cliente_id: uuid.UUID,
    db: Session = Depends(get_db),
    current: Usuario = Depends(get_current_user),
):
    """Classifica (processual/comercial/ruído) os e-mails já sincronizados
    antes desta feature existir, que ainda estão sem categoria."""
    from app.models.email_cliente import EmailCliente
    from app.services.ia_email_classificacao import classificar_lote

    pendentes = db.query(EmailCliente).filter(
        EmailCliente.cliente_id == cliente_id,
        EmailCliente.categoria.is_(None),
    ).all()
    if not pendentes:
        return {"classificados": 0}

    LOTE = 20
    total = 0
    for i in range(0, len(pendentes), LOTE):
        chunk = pendentes[i:i + LOTE]
        categorias = classificar_lote([
            {"remetente": e.remetente, "assunto": e.assunto, "snippet": e.snippet}
            for e in chunk
        ])
        for email, categoria in zip(chunk, categorias):
            email.categoria = categoria
        total += len(chunk)
    db.commit()
    return {"classificados": total}


# ── Emails ────────────────────────────────────────────────────────────────────

class EmailSyncRequest(BaseModel):
    # Uma conta ("master" | "<uid>" | "<uid>:extra") ou várias (grupo).
    conta_google: str = "master"
    contas: list[str] | None = None


class EmailPrivacidadeRequest(BaseModel):
    privado: bool


class EmailOut(BaseModel):
    id: uuid.UUID
    gmail_message_id: str
    conta_google: str
    conta_email: str | None
    remetente: str | None
    destinatarios: str | None
    assunto: str | None
    snippet: str | None
    thread_id: str | None
    data: str | None  # ISO datetime string
    lido: bool
    privado: bool = False
    privado_por: uuid.UUID | None = None
    created_at: str

    model_config = {"from_attributes": True}


@router.get("/{cliente_id}/emails", response_model=list[EmailOut])
def listar_emails(
    cliente_id: uuid.UUID,
    db: Session = Depends(get_db),
    current: Usuario = Depends(get_current_user),
):
    from app.models.email_cliente import EmailCliente
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    emails = (
        db.query(EmailCliente)
        .filter(EmailCliente.cliente_id == cliente_id)
        .order_by(EmailCliente.data.desc().nullslast())
        .all()
    )
    is_super = current.role == "super_admin"
    return [
        {
            "id": e.id,
            "gmail_message_id": e.gmail_message_id,
            "conta_google": e.conta_google,
            "conta_email": e.conta_email,
            "remetente": e.remetente,
            "destinatarios": e.destinatarios,
            "assunto": e.assunto,
            "snippet": e.snippet,
            "thread_id": e.thread_id,
            "data": e.data.isoformat() if e.data else None,
            "lido": e.lido,
            "privado": bool(e.privado),
            "privado_por": e.privado_por,
            "created_at": e.created_at.isoformat(),
        }
        for e in emails
        # Público por padrão; privado só aparece p/ quem marcou e super-admins.
        if (not e.privado) or is_super or (e.privado_por == current.id)
    ]


@router.post("/{cliente_id}/emails/sync")
def sincronizar_emails(
    cliente_id: uuid.UUID,
    body: EmailSyncRequest,
    db: Session = Depends(get_db),
    current: Usuario = Depends(get_current_user),
):
    from app.services.gmail_sync import sync_emails_for_client

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    # Endereços do cliente a procurar no Gmail
    emails_cliente: list[str] = []
    if cliente.email:
        emails_cliente.append(cliente.email)

    # Grupo de contas a varrer. Um usuário só pode usar 'master' ou as PRÓPRIAS
    # contas (principal/extra) — não as de outro usuário.
    contas = body.contas if body.contas else [body.conta_google]
    permitidas = {"master", str(current.id), f"{current.id}:extra"}
    contas = [c for c in contas if c in permitidas]
    if not contas:
        raise HTTPException(status_code=400, detail="Nenhuma conta válida selecionada.")

    total_synced = 0
    total_new = 0
    erros: list[str] = []
    for conta in contas:
        result = sync_emails_for_client(
            cliente_id=str(cliente_id),
            conta_google=conta,
            db_session=db,
            emails_cliente=emails_cliente,
        )
        total_synced += int(result.get("synced") or 0)
        total_new += int(result.get("new") or 0)
        if result.get("error"):
            erros.append(f"{conta}: {result['error']}")

    # Só falha de fato se TODAS as contas erraram.
    if erros and len(erros) == len(contas):
        raise HTTPException(status_code=400, detail=" | ".join(erros))

    return {"synced": total_synced, "new": total_new, "error": " | ".join(erros) or None}


@router.patch("/{cliente_id}/emails/{email_id}/privacidade")
def marcar_email_privacidade(
    cliente_id: uuid.UUID,
    email_id: uuid.UUID,
    body: EmailPrivacidadeRequest,
    db: Session = Depends(get_db),
    current: Usuario = Depends(get_current_user),
):
    from app.models.email_cliente import EmailCliente
    e = (
        db.query(EmailCliente)
        .filter(EmailCliente.id == email_id, EmailCliente.cliente_id == cliente_id)
        .first()
    )
    if not e:
        raise HTTPException(status_code=404, detail="Email não encontrado")
    is_super = current.role == "super_admin"
    # Tornar privado: qualquer usuário pode. Tornar público de novo: só quem
    # marcou ou um super-admin.
    if not body.privado and e.privado and not (is_super or e.privado_por == current.id):
        raise HTTPException(status_code=403, detail="Apenas quem marcou como privado pode reverter.")
    e.privado = body.privado
    e.privado_por = current.id if body.privado else None
    db.commit()
    return {"id": str(e.id), "privado": e.privado}


@router.post("/{cliente_id}/pasta-arquivos/refresh")
def refresh_pasta_arquivos(cliente_id: uuid.UUID, db: Session = Depends(get_db)):
    """Compat: returns persistent client Uploads from Google Drive."""
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    from app.services.google_drive import listar_arquivos
    return listar_arquivos(cliente.nome, "Uploads")
