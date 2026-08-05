"""
Exportação do inventário patrimonial de um cliente em XLS e PDF.

- XLS: planilha achatada (openpyxl), 1 linha por bem, cabeçalho estilizado.
- PDF: documento ReportLab Platypus com identidade visual do escritório
  (teal/dark), um "card" por bem envolto em KeepTogether para nunca quebrar
  um bem no meio de duas páginas, resumo no topo e rodapé com numeração.
"""
import io
from datetime import date
from pathlib import Path
from typing import Any

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Image,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

_LOGO_PATH = Path(__file__).parent.parent.parent / "logo.png"

# ── Paleta (mesma do frontend)
DARK = colors.HexColor("#1d1e20")
TEAL = colors.HexColor("#00b090")
TEAL_LIGHT = colors.HexColor("#e8f7f4")
LIGHT = colors.HexColor("#f2f3f6")
MID = colors.HexColor("#6a7070")
BORDER = colors.HexColor("#e2e4e8")
WHITE = colors.white

STATUS_LABEL = {"em_validacao": "Em validação", "validado": "Validado", "incerto": "Incerto"}
STATUS_BG = {
    "em_validacao": colors.HexColor("#fef3c7"),
    "validado": colors.HexColor("#d1fae5"),
    "incerto": colors.HexColor("#fee2e2"),
}
STATUS_FG = {
    "em_validacao": colors.HexColor("#92400e"),
    "validado": colors.HexColor("#065f46"),
    "incerto": colors.HexColor("#b91c1c"),
}
OBJETIVO_LABEL = {"venda": "Venda", "aluguel": "Aluguel", "segurar": "Segurar"}
TIPO_LABEL = {"movel": "Móvel", "imovel": "Imóvel"}
TIPO_DOC_LABEL = {
    "contrato_compra_venda": "Contrato de compra e venda",
    "escritura_publica": "Escritura pública",
    "cessao_direitos": "Cessão de direitos",
    "matricula": "Matrícula / Registro",
    "formal_partilha": "Formal de partilha",
    "outro": "Outro documento",
}
_MESES = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho",
          "agosto", "setembro", "outubro", "novembro", "dezembro"]


def _brl(v: Any) -> str:
    if v is None:
        return "—"
    try:
        s = f"{float(v):,.2f}"
    except (TypeError, ValueError):
        return "—"
    # 1,234,567.89 -> 1.234.567,89
    s = s.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {s}"


def _data(d: Any) -> str:
    if not d:
        return "—"
    try:
        return d.strftime("%d/%m/%Y")
    except AttributeError:
        return str(d)


def _norm(x: str | None) -> str:
    return " ".join((x or "").strip().lower().split())


# ── Ganho de capital (mesma lógica do frontend) ──────────────────────────────
def _num(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _meses_entre(inicio: date, fim: date) -> int:
    return max(0, (fim.year - inicio.year) * 12 + (fim.month - inicio.month))


def _irpf_progressivo(ganho: float) -> float:
    faixas = [(5_000_000, 0.15), (10_000_000, 0.175), (30_000_000, 0.20), (float("inf"), 0.225)]
    imposto = 0.0
    anterior = 0.0
    for teto, aliq in faixas:
        if ganho <= anterior:
            break
        imposto += (min(ganho, teto) - anterior) * aliq
        anterior = teto
    return imposto


def _fator_reducao_imovel(data_compra: date, data_venda: date) -> float:
    """Fração do ganho que permanece tributável na PF (Leis 11.196/2005 e 7.713/88)."""
    ano = data_compra.year
    if ano <= 1969:
        mult7713 = 0.0
    elif ano <= 1988:
        mult7713 = 5 * (ano - 1969) / 100
    else:
        mult7713 = 1.0
    nov2005, dez2005 = date(2005, 11, 1), date(2005, 12, 1)
    m1 = _meses_entre(data_compra, nov2005) if data_compra < nov2005 else 0
    inicio_f2 = data_compra if data_compra > dez2005 else dez2005
    m2 = _meses_entre(inicio_f2, data_venda)
    # FR1 = 0,60%/mês (aquisição→nov/2005); FR2 = 0,35%/mês (dez/2005→alienação) — Lei 11.196/2005
    return mult7713 * (1 / (1.006 ** m1)) * (1 / (1.0035 ** m2))


def _gc_imovel(b, hoje: date) -> dict:
    aquis = _num(b.valor_compra) or _num(b.valor_ir)
    venda = _num(b.valor_mercado)
    ganho = max(0.0, venda - aquis)
    fator = _fator_reducao_imovel(b.data_compra, hoje) if b.data_compra else 1.0
    imp_pf = _irpf_progressivo(ganho * fator)
    imp_pj = ganho * 0.34
    imp_hold = venda * 0.0673
    menor = min(imp_pf, imp_pj, imp_hold)
    return {"aquis": aquis, "venda": venda, "ganho": ganho, "fator": fator,
            "imp_pf": imp_pf, "imp_pj": imp_pj, "imp_hold": imp_hold, "menor": menor}


def _gc_cota(b) -> dict:
    custo = _num(b.valor_compra) or _num(b.capital_social) or _num(b.valor_ir)
    venda = _num(b.valor_mercado) or _num(b.valor_balanco)
    ganho = max(0.0, venda - custo)
    return {"custo": custo, "venda": venda, "ganho": ganho, "imp_pf": _irpf_progressivo(ganho)}


def _eh_cota(b) -> bool:
    return b.tipo_bem == "movel" and bool(
        b.empresa_nome or b.capital_social is not None or b.valor_balanco is not None or b.socios
    )


# ════════════════════════════════════════════════════════════════════════════
# XLS
# ════════════════════════════════════════════════════════════════════════════
def gerar_xls(cliente_nome: str, bens: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patrimônio"

    titulo = ws.cell(row=1, column=1, value=f"Inventário Patrimonial — {cliente_nome}")
    titulo.font = Font(bold=True, size=14, color="1D1E20")
    ws.cell(row=2, column=1,
            value=f"Gerado em {date.today().strftime('%d/%m/%Y')} · {len(bens)} bem(ns)").font = Font(
        italic=True, size=9, color="6A7070")

    colunas = [
        ("Tipo", 12), ("Nome do bem", 34), ("Descrição", 34), ("Objetivo", 12),
        ("Valor de compra", 16), ("Valor de mercado", 16), ("Valor no IR", 16),
        ("Data da compra", 14), ("Status", 14), ("Integralizar holding", 18),
        ("Nº matrícula", 16), ("Cartório", 24), ("Descrição matrícula", 34),
        ("Proprietário real", 24), ("Proprietário na matrícula", 26), ("Confere?", 12),
        ("Gravame?", 10), ("Detalhe gravame", 28), ("Observações", 34),
        ("Cadeia sucessória", 40), ("Anexos", 30),
    ]
    header_row = 4
    for idx, (nome, largura) in enumerate(colunas, start=1):
        c = ws.cell(row=header_row, column=idx, value=nome)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(start_color="00B090", end_color="00B090", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = largura

    def _money(v):
        return float(v) if v is not None else None

    total_mercado = total_ir = total_compra = 0.0
    r = header_row + 1
    for b in bens:
        real, matp = _norm(b.proprietario_real), _norm(b.proprietario_matricula)
        confere = "—"
        if real and matp:
            confere = "Sim" if real == matp else "Diverge"
        cadeia = " | ".join(
            f"{i+1}. {TIPO_DOC_LABEL.get(e.tipo_documento, e.tipo_documento)}"
            + (f" ({e.de_quem or '?'} → {e.para_quem or '?'})" if (e.de_quem or e.para_quem) else "")
            + (f" {_data(e.data)}" if e.data else "")
            for i, e in enumerate(sorted(b.cadeia, key=lambda x: x.ordem))
        )
        anexos = ", ".join(a.filename for a in b.anexos)
        valores = [
            TIPO_LABEL.get(b.tipo_bem, b.tipo_bem), b.nome, b.descricao or "",
            OBJETIVO_LABEL.get(b.objetivo, "") if b.objetivo else "",
            _money(b.valor_compra), _money(b.valor_mercado), _money(b.valor_ir),
            _data(b.data_compra), STATUS_LABEL.get(b.status, b.status),
            "Sim" if b.integralizar_holding else "Não",
            b.numero_matricula or "", b.cartorio or "", b.descricao_matricula or "",
            b.proprietario_real or "", b.proprietario_matricula or "", confere,
            "Sim" if b.tem_gravame else "Não", b.gravame_descricao or "",
            b.observacoes or "", cadeia, anexos,
        ]
        for idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=idx in (2, 3, 13, 19, 20, 21))
            if idx in (5, 6, 7) and val is not None:
                cell.number_format = 'R$ #,##0.00'
        total_compra += _money(b.valor_compra) or 0
        total_mercado += _money(b.valor_mercado) or 0
        total_ir += _money(b.valor_ir) or 0
        r += 1

    # Linha de totais
    tot = ws.cell(row=r, column=4, value="TOTAIS")
    tot.font = Font(bold=True, color="1D1E20")
    for idx, total in ((5, total_compra), (6, total_mercado), (7, total_ir)):
        c = ws.cell(row=r, column=idx, value=total)
        c.font = Font(bold=True, color="00B090")
        c.number_format = 'R$ #,##0.00'

    ws.freeze_panes = f"A{header_row + 1}"

    # ── Aba: Ganho de Capital — Imóveis ──────────────────────────────────────
    from datetime import date as _date
    hoje = _date.today()

    def _header_aba(ws2, titulo, colunas):
        ws2.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=13, color="1D1E20")
        for i, (nome, larg) in enumerate(colunas, start=1):
            c = ws2.cell(row=3, column=i, value=nome)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill(start_color="00B090", end_color="00B090", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws2.column_dimensions[get_column_letter(i)].width = larg

    imoveis = [b for b in bens if b.tipo_bem == "imovel"]
    if imoveis:
        wi = wb.create_sheet("GC Imóveis")
        _header_aba(wi, f"Ganho de Capital — Imóveis (venda estimada hoje, {hoje.strftime('%d/%m/%Y')})", [
            ("Imóvel", 34), ("Nº matrícula", 16), ("Aquisição", 16), ("Venda estimada", 16),
            ("Ganho", 16), ("Redução PF (base) %", 18), ("IR PF", 16), ("PJ 34%", 16), ("Holding 6,73%", 16),
        ])
        r = 4
        tot = {"ganho": 0.0, "pf": 0.0, "pj": 0.0, "hold": 0.0}
        for b in imoveis:
            g = _gc_imovel(b, hoje)
            red = (1 - g["fator"]) * 100 if b.data_compra else None
            vals = [b.nome, b.numero_matricula or "", g["aquis"], g["venda"], g["ganho"],
                    (red if red is not None else "s/ data"), g["imp_pf"], g["imp_pj"], g["imp_hold"]]
            for i, v in enumerate(vals, start=1):
                cell = wi.cell(row=r, column=i, value=v)
                if i in (3, 4, 5, 7, 8, 9):
                    cell.number_format = 'R$ #,##0.00'
                if i == 6 and isinstance(v, (int, float)):
                    cell.number_format = '0.0"%"'
            tot["ganho"] += g["ganho"]; tot["pf"] += g["imp_pf"]; tot["pj"] += g["imp_pj"]; tot["hold"] += g["imp_hold"]
            r += 1
        wi.cell(row=r, column=2, value="TOTAIS").font = Font(bold=True)
        for i, val in ((5, tot["ganho"]), (7, tot["pf"]), (8, tot["pj"]), (9, tot["hold"])):
            c = wi.cell(row=r, column=i, value=val)
            c.font = Font(bold=True, color="00B090"); c.number_format = 'R$ #,##0.00'
        wi.freeze_panes = "A4"

    # ── Aba: Ganho de Capital — Cotas ────────────────────────────────────────
    cotas = [b for b in bens if _eh_cota(b)]
    if cotas:
        wc = wb.create_sheet("GC Cotas")
        _header_aba(wc, "Ganho de Capital — Cotas/Participações (PF 15–22,5%, sem fator de redução)", [
            ("Participação", 34), ("CNPJ", 20), ("Custo/capital", 16), ("Valor estimado", 16),
            ("Ganho", 16), ("IR PF", 16),
        ])
        r = 4
        tot = {"ganho": 0.0, "pf": 0.0}
        for b in cotas:
            g = _gc_cota(b)
            vals = [b.empresa_nome or b.nome, b.empresa_cnpj or "", g["custo"], g["venda"], g["ganho"], g["imp_pf"]]
            for i, v in enumerate(vals, start=1):
                cell = wc.cell(row=r, column=i, value=v)
                if i in (3, 4, 5, 6):
                    cell.number_format = 'R$ #,##0.00'
            tot["ganho"] += g["ganho"]; tot["pf"] += g["imp_pf"]
            r += 1
        wc.cell(row=r, column=2, value="TOTAIS").font = Font(bold=True)
        for i, val in ((5, tot["ganho"]), (6, tot["pf"])):
            c = wc.cell(row=r, column=i, value=val)
            c.font = Font(bold=True, color="00B090"); c.number_format = 'R$ #,##0.00'
        wc.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# PDF
# ════════════════════════════════════════════════════════════════════════════
def _estilos():
    return {
        "titulo": ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=17,
                                 textColor=DARK, leading=20, spaceAfter=5, alignment=TA_LEFT),
        "sub": ParagraphStyle("sub", fontName="Helvetica", fontSize=10, textColor=MID,
                              leading=13, spaceAfter=2),
        "kpiVal": ParagraphStyle("kpiVal", fontName="Helvetica-Bold", fontSize=13,
                                 textColor=DARK, alignment=TA_CENTER, leading=15),
        "kpiValTeal": ParagraphStyle("kpiValTeal", fontName="Helvetica-Bold", fontSize=13,
                                     textColor=TEAL, alignment=TA_CENTER, leading=15),
        "kpiLbl": ParagraphStyle("kpiLbl", fontName="Helvetica", fontSize=7.5, textColor=MID,
                                 alignment=TA_CENTER, spaceBefore=2),
        "bemNome": ParagraphStyle("bemNome", fontName="Helvetica-Bold", fontSize=11.5,
                                  textColor=WHITE, leading=14),
        "statusPill": ParagraphStyle("statusPill", fontName="Helvetica-Bold", fontSize=8,
                                     alignment=TA_CENTER),
        "lbl": ParagraphStyle("lbl", fontName="Helvetica-Bold", fontSize=7.5, textColor=MID),
        "val": ParagraphStyle("val", fontName="Helvetica", fontSize=9.5, textColor=DARK,
                              leading=12),
        "valStrong": ParagraphStyle("valStrong", fontName="Helvetica-Bold", fontSize=9.5,
                                    textColor=DARK, leading=12),
        "secao": ParagraphStyle("secao", fontName="Helvetica-Bold", fontSize=8.5, textColor=TEAL,
                                spaceBefore=2, spaceAfter=2),
        "elo": ParagraphStyle("elo", fontName="Helvetica", fontSize=9, textColor=DARK,
                              leading=13, leftIndent=6),
        "matchOk": ParagraphStyle("matchOk", fontName="Helvetica-Bold", fontSize=8.5,
                                  textColor=colors.HexColor("#065f46")),
        "matchDiff": ParagraphStyle("matchDiff", fontName="Helvetica-Bold", fontSize=8.5,
                                    textColor=colors.HexColor("#92400e")),
        "obs": ParagraphStyle("obs", fontName="Helvetica-Oblique", fontSize=9, textColor=MID,
                              leading=12),
    }


def _rodape(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(colors.HexColor("#9ca3af"))
    canvas.drawString(1.8 * cm, 1.1 * cm, "Pimenta Júdice Advogados · Inventário Patrimonial")
    canvas.drawRightString(A4[0] - 1.8 * cm, 1.1 * cm, f"Página {doc.page}")
    canvas.setStrokeColor(BORDER)
    canvas.setLineWidth(0.5)
    canvas.line(1.8 * cm, 1.35 * cm, A4[0] - 1.8 * cm, 1.35 * cm)
    canvas.restoreState()


def _card_bem(b, st, content_w) -> KeepTogether:
    els: list = []

    # ── Barra de título (dark) com nome + status pill
    status_pill = Table(
        [[Paragraph(STATUS_LABEL.get(b.status, b.status),
                    ParagraphStyle("sp", parent=st["statusPill"], textColor=STATUS_FG.get(b.status, MID)))]],
        colWidths=[3.2 * cm],
    )
    status_pill.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), STATUS_BG.get(b.status, LIGHT)),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ROUNDEDCORNERS", [4, 4, 4, 4]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    icone = "IMÓVEL" if b.tipo_bem == "imovel" else "MÓVEL"
    tags = []
    if b.integralizar_holding:
        tags.append("HOLDING")
    if b.tem_gravame:
        tags.append("GRAVAME")
    nome_txt = b.nome + (f'  <font size="7" color="#a7f3d0">[{" · ".join(tags)}]</font>' if tags else "")
    barra = Table(
        [[Paragraph(f'<font size="7" color="#7fd8c6">{icone}</font>  {nome_txt}', st["bemNome"]),
          status_pill]],
        colWidths=[content_w - 3.4 * cm, 3.4 * cm],
    )
    barra.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), DARK),
        ("BACKGROUND", (1, 0), (1, -1), DARK),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, 0), 10), ("RIGHTPADDING", (1, 0), (1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    els.append(barra)

    # ── Grade de campos (label/valor em 2 pares por linha)
    def cell(label, value, strong=False):
        return [Paragraph(label.upper(), st["lbl"]),
                Paragraph(value, st["valStrong"] if strong else st["val"])]

    pares = [
        ("Tipo", TIPO_LABEL.get(b.tipo_bem, b.tipo_bem)),
        ("Objetivo", OBJETIVO_LABEL.get(b.objetivo, "—") if b.objetivo else "—"),
        ("Valor de compra", _brl(b.valor_compra)),
        ("Valor de mercado", _brl(b.valor_mercado)),
        ("Valor no IR", _brl(b.valor_ir)),
        ("Data da compra", _data(b.data_compra)),
    ]
    if b.tipo_bem == "imovel":
        pares += [("Nº matrícula", b.numero_matricula or "—"),
                  ("Cartório", b.cartorio or "—")]

    col_lbl = 2.6 * cm
    col_val = (content_w - 2 * col_lbl) / 2
    rows = []
    for i in range(0, len(pares), 2):
        left = cell(*pares[i], strong=("Valor" in pares[i][0]))
        if i + 1 < len(pares):
            right = cell(*pares[i + 1], strong=("Valor" in pares[i + 1][0]))
        else:
            right = [Paragraph("", st["lbl"]), Paragraph("", st["val"])]
        rows.append([left[0], left[1], right[0], right[1]])
    grade = Table(rows, colWidths=[col_lbl, col_val, col_lbl, col_val])
    grade.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, BORDER),
        ("BACKGROUND", (0, 0), (-1, -1), WHITE),
        ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
    ]))
    els.append(grade)

    # ── Descrição do bem
    if b.descricao:
        els.append(_faixa([Paragraph("DESCRIÇÃO", st["lbl"]),
                           Paragraph(b.descricao, st["val"])], content_w))

    # ── Proprietários + match
    real, matp = _norm(b.proprietario_real), _norm(b.proprietario_matricula)
    prop_rows = [[
        Paragraph("PROPRIETÁRIO REAL", st["lbl"]),
        Paragraph("PROPRIETÁRIO NA MATRÍCULA", st["lbl"]),
        Paragraph("", st["lbl"]),
    ], [
        Paragraph(b.proprietario_real or "—", st["val"]),
        Paragraph(b.proprietario_matricula or "—", st["val"]),
        (Paragraph("✓ Confere", st["matchOk"]) if real and matp and real == matp
         else Paragraph("⚠ Diverge", st["matchDiff"]) if real and matp
         else Paragraph("", st["val"])),
    ]]
    prop = Table(prop_rows, colWidths=[content_w * 0.4, content_w * 0.4, content_w * 0.2])
    prop.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
    ]))
    els.append(prop)

    # ── Gravame
    if b.tem_gravame:
        els.append(_faixa([Paragraph("GRAVAME / ÔNUS", st["lbl"]),
                           Paragraph(b.gravame_descricao or "Sim (sem detalhes)", st["val"])],
                          content_w, bg=colors.HexColor("#fef2f2")))

    # ── Descrição conforme matrícula
    if b.descricao_matricula:
        els.append(_faixa([Paragraph("DESCRIÇÃO CONFORME MATRÍCULA", st["lbl"]),
                           Paragraph(b.descricao_matricula, st["val"])], content_w))

    # ── Observações
    if b.observacoes:
        els.append(_faixa([Paragraph("OBSERVAÇÕES", st["lbl"]),
                           Paragraph(b.observacoes, st["obs"])], content_w))

    # ── Cadeia sucessória
    if b.tipo_bem == "imovel" and b.cadeia:
        cadeia_els = [Paragraph("CADEIA SUCESSÓRIA", st["secao"])]
        for i, e in enumerate(sorted(b.cadeia, key=lambda x: x.ordem), start=1):
            partes = f"{e.de_quem or '—'} → {e.para_quem or '—'}" if (e.de_quem or e.para_quem) else ""
            extra = " · ".join(x for x in [partes, _data(e.data) if e.data else ""] if x)
            txt = f"<b>{i}.</b> {TIPO_DOC_LABEL.get(e.tipo_documento, e.tipo_documento)}"
            if extra:
                txt += f' <font color="#6a7070">— {extra}</font>'
            if e.descricao:
                txt += f'<br/><font size="8" color="#6a7070">{e.descricao}</font>'
            cadeia_els.append(Paragraph(txt, st["elo"]))
        wrap = Table([[c] for c in cadeia_els], colWidths=[content_w])
        wrap.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), TEAL_LIGHT),
            ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (0, 0), 6), ("BOTTOMPADDING", (0, -1), (-1, -1), 6),
            ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
        ]))
        els.append(wrap)

    els.append(Spacer(1, 0.45 * cm))
    return KeepTogether(els)


def _faixa(paras, content_w, bg=LIGHT):
    t = Table([[p] for p in paras], colWidths=[content_w])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), bg),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (0, 0), 5), ("BOTTOMPADDING", (0, -1), (-1, -1), 5),
        ("TOPPADDING", (0, 1), (-1, -1), 1),
        ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
    ]))
    return t


def _gc_pdf(bens, st, content_w) -> list:
    """Seções de Ganho de Capital (imóveis + cotas) para o PDF."""
    from datetime import date as _date
    hoje = _date.today()
    imoveis = [b for b in bens if b.tipo_bem == "imovel"]
    cotas = [b for b in bens if _eh_cota(b)]
    els: list = []
    if not imoveis and not cotas:
        return els

    header_st = ParagraphStyle("gchead", fontName="Helvetica-Bold", fontSize=7.5, textColor=WHITE)
    cel_st = ParagraphStyle("gccel", fontName="Helvetica", fontSize=7.5, textColor=DARK, leading=9)
    tot_st = ParagraphStyle("gctot", fontName="Helvetica-Bold", fontSize=7.5, textColor=DARK)

    def _p(txt, style=cel_st):
        return Paragraph(str(txt), style)

    els.append(Paragraph("Análise de Ganho de Capital", st["secao"]))
    els.append(Paragraph(
        f"Venda estimada hoje ({hoje.strftime('%d/%m/%Y')}). Na PF, o fator de redução incide sobre a "
        f"base (o ganho): Leis 11.196/2005 (FR1 0,60%/mês e FR2 0,35%/mês) e 7.713/88 (imóveis até 1969 isentos; 1970–1988 decrescente).",
        st["obs"]))

    # ── Imóveis
    if imoveis:
        els.append(Paragraph("Imóveis", st["lbl"]))
        head = ["Imóvel", "Aquisição", "Venda est.", "Ganho", "Red. PF", "IR PF", "PJ 34%", "Holding"]
        rows = [[_p(h, header_st) for h in head]]
        menor_cols = []
        tot = {"g": 0.0, "pf": 0.0, "pj": 0.0, "hold": 0.0}
        for b in imoveis:
            g = _gc_imovel(b, hoje)
            red = f"−{(1 - g['fator']) * 100:.1f}%".replace(".", ",") if b.data_compra else "s/ data"
            rows.append([
                _p(b.nome), _p(_brl(g["aquis"])), _p(_brl(g["venda"])), _p(_brl(g["ganho"])),
                _p(red), _p(_brl(g["imp_pf"])), _p(_brl(g["imp_pj"])), _p(_brl(g["imp_hold"])),
            ])
            menor_cols.append(5 if g["menor"] == g["imp_pf"] else 6 if g["menor"] == g["imp_pj"] else 7)
            tot["g"] += g["ganho"]; tot["pf"] += g["imp_pf"]; tot["pj"] += g["imp_pj"]; tot["hold"] += g["imp_hold"]
        rows.append([_p("TOTAIS", tot_st), _p(""), _p(""), _p(_brl(tot["g"]), tot_st), _p(""),
                     _p(_brl(tot["pf"]), tot_st), _p(_brl(tot["pj"]), tot_st), _p(_brl(tot["hold"]), tot_st)])
        cw = [c * content_w for c in (0.224, 0.121, 0.121, 0.121, 0.086, 0.121, 0.103, 0.103)]
        t = Table(rows, colWidths=cw, repeatRows=1)
        ts = [
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("BACKGROUND", (0, -1), (-1, -1), LIGHT),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("LINEBELOW", (0, 0), (-1, -2), 0.4, BORDER),
            ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
        ]
        for i, col in enumerate(menor_cols, start=1):
            ts.append(("BACKGROUND", (col, i), (col, i), TEAL_LIGHT))
        t.setStyle(TableStyle(ts))
        els.append(t)
        els.append(Spacer(1, 0.3 * cm))

    # ── Cotas
    if cotas:
        els.append(Paragraph("Cotas / participações societárias (PF 15–22,5%, sem fator de redução)", st["lbl"]))
        head = ["Participação", "CNPJ", "Custo/capital", "Valor est.", "Ganho", "IR PF"]
        rows = [[_p(h, header_st) for h in head]]
        tot = {"g": 0.0, "pf": 0.0}
        for b in cotas:
            g = _gc_cota(b)
            rows.append([
                _p(b.empresa_nome or b.nome), _p(b.empresa_cnpj or "—"),
                _p(_brl(g["custo"])), _p(_brl(g["venda"])), _p(_brl(g["ganho"])), _p(_brl(g["imp_pf"])),
            ])
            tot["g"] += g["ganho"]; tot["pf"] += g["imp_pf"]
        rows.append([_p("TOTAIS", tot_st), _p(""), _p(""), _p(""), _p(_brl(tot["g"]), tot_st), _p(_brl(tot["pf"]), tot_st)])
        cw = [c * content_w for c in (0.29, 0.19, 0.13, 0.13, 0.13, 0.13)]
        t = Table(rows, colWidths=cw, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("BACKGROUND", (0, -1), (-1, -1), LIGHT),
            ("BACKGROUND", (5, 1), (5, -2), TEAL_LIGHT),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("LINEBELOW", (0, 0), (-1, -2), 0.4, BORDER),
            ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
        ]))
        els.append(t)

    els.append(Spacer(1, 0.5 * cm))
    return els


def gerar_pdf(cliente_nome: str, bens: list) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
        topMargin=1.6 * cm, bottomMargin=1.8 * cm,
        title=f"Inventário Patrimonial — {cliente_nome}",
    )
    content_w = A4[0] - 3.6 * cm
    st = _estilos()
    story: list = []

    # ── Cabeçalho
    if _LOGO_PATH.exists():
        img = Image(str(_LOGO_PATH))
        max_h = 1.7 * cm
        ratio = max_h / img.imageHeight
        img.drawWidth = img.imageWidth * ratio
        img.drawHeight = max_h
        img.hAlign = "LEFT"
        story.append(img)
        story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph("Inventário Patrimonial", st["titulo"]))
    hoje = date.today()
    story.append(Paragraph(
        f"{cliente_nome} · gerado em {hoje.day} de {_MESES[hoje.month]} de {hoje.year}", st["sub"]))
    story.append(HRFlowable(width="100%", thickness=1, color=TEAL, spaceBefore=6, spaceAfter=10))

    # ── Resumo (KPIs)
    total_mercado = sum(float(b.valor_mercado or 0) for b in bens)
    total_ir = sum(float(b.valor_ir or 0) for b in bens)
    n_holding = sum(1 for b in bens if b.integralizar_holding)

    def kpi(valor, label, teal=False):
        return [Paragraph(valor, st["kpiValTeal"] if teal else st["kpiVal"]),
                Paragraph(label, st["kpiLbl"])]

    kpi_cells = [
        kpi(str(len(bens)), "BENS CADASTRADOS"),
        kpi(_brl(total_mercado), "VALOR DE MERCADO", teal=True),
        kpi(_brl(total_ir), "VALOR NO IR"),
        kpi(str(n_holding), "PARA A HOLDING"),
    ]
    resumo = Table([[c[0] for c in kpi_cells], [c[1] for c in kpi_cells]],
                   colWidths=[content_w / 4] * 4)
    resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 9), ("BOTTOMPADDING", (0, 1), (-1, 1), 9),
        ("LINEAFTER", (0, 0), (-2, -1), 0.6, BORDER),
        ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
    ]))
    story.append(resumo)
    story.append(Spacer(1, 0.5 * cm))

    # Análise de Ganho de Capital (imóveis + cotas)
    for el in _gc_pdf(bens, st, content_w):
        story.append(el)

    if not bens:
        story.append(Paragraph("Nenhum bem cadastrado.", st["val"]))
    else:
        for b in bens:
            story.append(_card_bem(b, st, content_w))

    doc.build(story, onFirstPage=_rodape, onLaterPages=_rodape)
    return buf.getvalue()
